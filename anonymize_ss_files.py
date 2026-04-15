"""
Anonymize ss_files/ → ss_files_anon/

Rules:
- Keep only accounts that appear as subject_id in transactions
- Keep only customers linked to those active accounts (via account_relationship)
- Replace account UUIDs  → acct_id_XXXXX  (consistent across all files)
- Replace customer UUIDs → cust_id_XXXXX  (consistent across all files)
- Replace alert IDs      → alert_id_XXXXX (consistent across all files)
- Leave other_party_id, conductor_id, and non-ID columns unchanged
"""

import csv
import re
import os
import json
import openpyxl
from copy import copy

SRC = r"C:\Users\Aaditya\PycharmProjects\framlagents-ftllm\ss_files"
DST = r"C:\Users\Aaditya\PycharmProjects\framlagents-ftllm\ss_files_anon"
os.makedirs(DST, exist_ok=True)

# ── helpers ────────────────────────────────────────────────────────────────────

UUID_RE = re.compile(
    r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", re.IGNORECASE
)

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader), reader.fieldnames

def write_csv(path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)

def make_id_map(uuids, prefix):
    """Build {uuid: prefix_XXXXX} mapping, sorted for determinism."""
    return {u: f"{prefix}_{i+1:05d}" for i, u in enumerate(sorted(uuids))}

def replace_set_literal(cell_value, acct_map, cust_map):
    """Replace UUIDs inside postgres set literal  {uuid1,uuid2,...}  or JSON set."""
    if not cell_value:
        return cell_value
    s = str(cell_value)
    def _sub(m):
        uid = m.group(0).lower()
        return cust_map.get(uid, acct_map.get(uid, m.group(0)))
    return UUID_RE.sub(_sub, s)

# ── step 1: find active account IDs from transactions ─────────────────────────

print("[1] Reading transactions to find active account IDs...")
trx_rows, trx_fields = read_csv(os.path.join(SRC, "aml_s_transactions.csv"))
active_account_uuids = {r["subject_id"].lower() for r in trx_rows if r.get("subject_id")}
print(f"    Active accounts in transactions: {len(active_account_uuids)}")

# ── step 2: find customer IDs linked to active accounts ───────────────────────

print("[2] Reading account_relationship to find linked customers...")
rel_rows, _ = read_csv(os.path.join(SRC, "aml_s_account_relationship.csv"))
active_rel_rows = [r for r in rel_rows if r["account_id"].lower() in active_account_uuids]
active_customer_uuids = {r["customer_id"].lower() for r in active_rel_rows}
print(f"    Linked customers: {len(active_customer_uuids)}")

# ── step 3: build ID maps ─────────────────────────────────────────────────────

print("[3] Building ID maps...")
acct_map = make_id_map(active_account_uuids, "acct_id")
cust_map = make_id_map(active_customer_uuids, "cust_id")

# Alert IDs — collect from both CSV and Excel first pass
alert_ids_raw = set()

def collect_alert_ids_csv(path):
    rows, _ = read_csv(path)
    for r in rows:
        v = r.get("Alert ID", "")
        if v:
            alert_ids_raw.add(v.strip())

collect_alert_ids_csv(os.path.join(SRC, "PSG_Alert_Report_env2_Nov3_condition_details.csv"))

wb_src = openpyxl.load_workbook(
    os.path.join(SRC, "PSG_Alert_Report_env2_Nov3_PSG_Alerts_11112025.xlsx")
)
main_ws = wb_src["PSG_Alert_Report_env2_Nov3"]
main_headers = [c.value for c in next(main_ws.iter_rows(min_row=1, max_row=1))]
alert_id_col = main_headers.index("Alert ID") if "Alert ID" in main_headers else None
if alert_id_col is not None:
    for row in main_ws.iter_rows(min_row=2, values_only=True):
        v = row[alert_id_col]
        if v:
            alert_ids_raw.add(str(v).strip())

alert_map = make_id_map(sorted(alert_ids_raw), "alert_id")
print(f"    Unique alert IDs: {len(alert_map)}")

# ── step 4: anonymize aml_s_accounts.csv ──────────────────────────────────────

print("[4] Anonymizing aml_s_accounts.csv...")
acct_rows, acct_fields = read_csv(os.path.join(SRC, "aml_s_accounts.csv"))
out_acct = []
for r in acct_rows:
    uid = r["id"].lower()
    if uid not in acct_map:
        continue  # skip inactive accounts
    r2 = dict(r)
    r2["id"] = acct_map[uid]
    out_acct.append(r2)
write_csv(os.path.join(DST, "aml_s_accounts.csv"), out_acct, acct_fields)
print(f"    Written {len(out_acct)} active accounts")

# ── step 5: anonymize aml_s_customers.csv ────────────────────────────────────

print("[5] Anonymizing aml_s_customers.csv...")
cust_rows, cust_fields = read_csv(os.path.join(SRC, "aml_s_customers.csv"))
out_cust = []
for r in cust_rows:
    uid = r["id"].lower()
    if uid not in cust_map:
        continue  # skip customers not linked to active accounts
    r2 = dict(r)
    r2["id"] = cust_map[uid]
    out_cust.append(r2)
write_csv(os.path.join(DST, "aml_s_customers.csv"), out_cust, cust_fields)
print(f"    Written {len(out_cust)} customers")

# ── step 6: anonymize aml_s_account_relationship.csv ─────────────────────────

print("[6] Anonymizing aml_s_account_relationship.csv...")
out_rel = []
for r in active_rel_rows:
    aid = r["account_id"].lower()
    cid = r["customer_id"].lower()
    if aid not in acct_map or cid not in cust_map:
        continue
    r2 = dict(r)
    r2["account_id"] = acct_map[aid]
    r2["customer_id"] = cust_map[cid]
    out_rel.append(r2)
_, rel_fields = read_csv(os.path.join(SRC, "aml_s_account_relationship.csv"))
write_csv(os.path.join(DST, "aml_s_account_relationship.csv"), out_rel, rel_fields)
print(f"    Written {len(out_rel)} relationship rows")

# ── step 7: anonymize aml_s_transactions.csv ─────────────────────────────────

print("[7] Anonymizing aml_s_transactions.csv...")
out_trx = []
for r in trx_rows:
    aid = r.get("subject_id", "").lower()
    if aid not in acct_map:
        continue
    r2 = dict(r)
    r2["subject_id"] = acct_map[aid]
    # customer_ids is a postgres set literal: {uuid1,uuid2} — replace each UUID
    r2["customer_ids"] = replace_set_literal(r.get("customer_ids", ""), acct_map, cust_map)
    # leave other_party_id and conductor_id as-is
    out_trx.append(r2)
write_csv(os.path.join(DST, "aml_s_transactions.csv"), out_trx, trx_fields)
print(f"    Written {len(out_trx)} transaction rows")

# ── step 8: anonymize condition_details.csv ───────────────────────────────────

print("[8] Anonymizing PSG_Alert_Report_env2_Nov3_condition_details.csv...")
cond_rows, cond_fields = read_csv(
    os.path.join(SRC, "PSG_Alert_Report_env2_Nov3_condition_details.csv")
)
out_cond = []
for r in cond_rows:
    r2 = dict(r)
    # Alert ID
    v = r2.get("Alert ID", "").strip()
    r2["Alert ID"] = alert_map.get(v, v)
    # Account Numbers = one or more comma-separated account UUIDs
    def replace_account_list(cell):
        parts = [p.strip() for p in str(cell).split(",")]
        return ",".join(acct_map.get(p.lower(), p) for p in parts)
    r2["Account Numbers"] = replace_account_list(r2.get("Account Numbers", ""))
    # Subject Account Number = account UUID
    san = r2.get("Subject Account Number", "").strip().lower()
    r2["Subject Account Number"] = acct_map.get(san, r2["Subject Account Number"])
    out_cond.append(r2)
write_csv(
    os.path.join(DST, "PSG_Alert_Report_env2_Nov3_condition_details.csv"),
    out_cond,
    cond_fields,
)
print(f"    Written {len(out_cond)} condition detail rows")

# ── step 9: anonymize Excel ───────────────────────────────────────────────────

print("[9] Anonymizing PSG_Alert_Report_env2_Nov3_PSG_Alerts_11112025.xlsx...")

wb_dst = openpyxl.load_workbook(
    os.path.join(SRC, "PSG_Alert_Report_env2_Nov3_PSG_Alerts_11112025.xlsx")
)

# --- Main sheet: PSG_Alert_Report_env2_Nov3 ---
ws = wb_dst["PSG_Alert_Report_env2_Nov3"]
hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

def col_idx(name):
    return hdrs.index(name) + 1 if name in hdrs else None

c_alert  = col_idx("Alert ID")
c_acct   = col_idx("Account Numbers")
c_subj   = col_idx("Subject Account Number")

for row in ws.iter_rows(min_row=2):
    if c_alert:
        v = str(row[c_alert-1].value or "").strip()
        row[c_alert-1].value = alert_map.get(v, v)
    if c_acct:
        parts = [p.strip() for p in str(row[c_acct-1].value or "").split(",")]
        row[c_acct-1].value = ",".join(acct_map.get(p.lower(), p) for p in parts)
    if c_subj:
        v = str(row[c_subj-1].value or "").strip().lower()
        row[c_subj-1].value = acct_map.get(v, row[c_subj-1].value)

# --- AD / EA / Burst sheets: subject_id + customer_ids ---
for sh_name in ["AD", "EA", "Burst"]:
    ws2 = wb_dst[sh_name]
    hdrs2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    c_sid = hdrs2.index("subject_id") + 1 if "subject_id" in hdrs2 else None
    c_cids = hdrs2.index("customer_ids") + 1 if "customer_ids" in hdrs2 else None
    for row in ws2.iter_rows(min_row=2):
        if c_sid:
            v = str(row[c_sid-1].value or "").strip().lower()
            row[c_sid-1].value = acct_map.get(v, row[c_sid-1].value)
        if c_cids:
            v = str(row[c_cids-1].value or "")
            row[c_cids-1].value = replace_set_literal(v, acct_map, cust_map)

wb_dst.save(
    os.path.join(DST, "PSG_Alert_Report_env2_Nov3_PSG_Alerts_11112025.xlsx")
)
print("    Excel written.")

# ── done ──────────────────────────────────────────────────────────────────────
print("\n[done] All files written to", DST)
print(f"  Accounts: {len(acct_map)}, Customers: {len(cust_map)}, Alerts: {len(alert_map)}")
